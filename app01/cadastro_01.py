# -*- coding: utf-8 -*-

import openpyxl, pprint
import os
import sys
import datetime
from openpyxl.styles import NamedStyle
from .models import Evento,Grupo_funcoes,Funcao,Grupo_eventos
from . import funcoes_banco


def grupo_eventos(planilha,empresa,current_user):

    wb = openpyxl.load_workbook(planilha)
    sheets = wb.sheetnames

    sheet0 = sheets[0]
    sheet = wb.get_sheet_by_name(sheet0)


    row=2
    colA = sheet['A' + str(row)].value
    colB = sheet['B' + str(row)].value
    

    while row<sheet.max_row+1 and colA=='EVENTO' and colB==empresa:
        colD = sheet['D' + str(row)].value
        lista_desc_eventos=[]
        while row<sheet.max_row+1 and colA=='EVENTO' and colB==empresa and colD==sheet['D' + str(row)].value:

            evento = sheet['C' + str(row)].value # Id do tipo da decisao
            evento_p = sheet['D' + str(row)].value # Id da decisao

            evento   = evento.strip()
            evento_p = evento_p.strip()

            
            if evento != evento_p:
                if evento not in lista_desc_eventos:
                    lista_desc_eventos.append(evento)


            colA = sheet['A' + str(row)].value
            colB = sheet['B' + str(row)].value
            colD = sheet['D' + str(row)].value

            row+=1
        if len(lista_1)>0:
            grupo_eventos_2(lista_desc_eventos,evento_p,empresa,current_user)

    return 1



def grupo_eventos_2(lista_desc_eventos,evento_p,empresa,current_user):

    ev_p=Evento.objects.filter(empresa=empresa,evento=evento_p).first()
    if ev_p is None:
        ev_p=Evento(
            empresa=empresa,
            tipo='V',
            evento=evento_p,
            exibe_excel=1,
            cancelado='N',
            cl_orcamentaria='O',
            ordenacao=0
        )
        ev_p.save()
        ev_p=Evento.objects.filter(empresa=empresa,evento=evento_p).first()
        id_evento=ev_p.id_evento
    else:
        id_evento=ev_p.id_evento


    lista_eventos = [grupo.evento_original for grupo in Grupo_eventos.objects.filter(empresa='SS')]
    lista_novos=[]


    for kk in range(len(lista_desc_eventos)):
        if lista_desc_eventos[kk] not in lista_eventos:
            evs=Grupo_eventos(
                empresa=empresa,
                evento_original=lista_desc_eventos[kk],
                evento_principal=evento_p
                )
            lista_novos.append(evs)
    if len(lista_novos)>0:
        Grupo_eventos.objects.bulk_create(lista_novos)


    lista_id_municipios = [mn.id_municipio for mn in Municipio.objects.filter(empresa='SS')]
    lista_id_eventos= [ev.id_evento for ev in Evento.objecst.filter(empresa=empresa, evento__in=lista_desc_eventos)]


    if len(lista_id_eventos)>0:
        for kk in range(len(lista_id_municipios)):
            query = Folhaevento.objects.filter(id_municipio=lista_id_municipios[kk],id_evento__in=lista_id_eventos)
            if len(query)>0:
                query.update(id_evento=id_evento)

